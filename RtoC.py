import re
import openpyxl

with open('email.txt','r') as f: 
    txt = f.read()
    # print(txt)
# txt = input()
# ^\d\. | ^\([a-z]\) | 
output = re.split("\n",txt)
output = [i for i in output if i != '']
# print(output)

# splits the para from the commenbt
for index, item in enumerate(output):
    output[index] = item.split(' - ')

# removes header from para
for index, item in enumerate(output):
    # print(type(re.sub('\d\. |\([a-z]+\)','',item[0]).lstrip(" ")))
    output[index][0] = re.sub('\d\. |\([a-z]+\)','',item[0]).lstrip(" ")

# if bottom items does not have section then add the previous section to it
for index, item in enumerate(output):
    if len(item) != 2: 
        # output[index - 1 ][0]
        output[index] = [' - ', item[0]]

# output the text
with open('output.txt','w') as o: 
    for para in output: 
        o.write('\n')
        # print(len(para))
        for item in para:
            o.write(item + '\n')

# output to excel
my_wb = openpyxl.Workbook()
my_sheet = my_wb.active


# sets the font properties
# my_sheet
# my_sheet.style.font.name = 'Arial'
# my_sheet.style.font.size = 8

my_sheet.merge_cells('A1:C1') 
my_sheet.cell(row = 1, column = 1 ).value = 'Departmental Comment'
my_sheet.cell(row = 1, column = 4 ).value = 'Response to Comment'

my_sheet.merge_cells('A2:C2') 
my_sheet.cell(row = 2, column = 1 ).value = 'From: (By Email)\nContact: \nDate: 2023'
my_sheet.cell(row = 2, column = 4 ).value = ""

my_sheet.merge_cells('A3:C3') 
my_sheet.cell(row = 3, column = 1 ).value = 'Departmental Comment'
my_sheet.cell(row = 3, column = 4 ).value = ""

adjustment = 4

my_sheet.cell(row = adjustment, column = 1 ).value = 'Item'
my_sheet.cell(row = adjustment, column = 2 ).value = 'Section'
my_sheet.cell(row = adjustment, column = 3 ).value = 'Comments'
my_sheet.cell(row = adjustment, column = 4 ).value = ""

for index, _ in enumerate(output):
    if index != 0:
        my_sheet.cell(row = index + adjustment + 1, column = 1 ).value = index
    my_sheet.cell(row = index + adjustment + 1, column = 2 ).value = output[index][0]
    my_sheet.cell(row = index + adjustment + 1, column = 3 ).value = output[index][1]
    my_sheet.cell(row = index + adjustment + 1, column = 4 ).value = ''
my_wb.save('output.xlsx')

