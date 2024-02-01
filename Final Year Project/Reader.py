# 0 Bound, 1 Date + Time, 2 Seq No, 3 Lane, 4 Speed, 5 Class, 6 No of Axle, [ Axle Weight, Axle Spacing]
import sys
from datetime import datetime
import openpyxl
import seaborn as sns
import matplotlib.pyplot as plt
import pandas as pd 
import numpy as np
from scipy.interpolate import UnivariateSpline
from matplotlib import pyplot as plt
import statistics
import scipy.stats 

class traffic_data:
      def __init__(traffic_data, Bound, Date_and_Time, Seq_No, Lane, Speed, Class, No_of_Axle, list_of_axle_weight_and_spacing):
        traffic_data.Bound = Bound
        traffic_data.Date_and_Time = Date_and_Time
        traffic_data.Seq_No = Seq_No
        traffic_data.Lane = Lane
        traffic_data.Speed = Speed
        traffic_data.Class = Class
        traffic_data.No_of_Axle = No_of_Axle
        traffic_data.list_of_axle_weight_and_spacing = list_of_axle_weight_and_spacing


#   def myfunc(abc):
#     print("Hello my name is " + abc.name)

# p1 = Person("John", 36)
# p1.myfunc()
def find_time(entry:list,other_entry:list):
    # date_and_time = date_and_time[1][-8:]
    date_obj1 = datetime.strptime(entry[1][-8:], '%H:%M:%S')
    date_obj2 = datetime.strptime(other_entry[1][-8:], '%H:%M:%S')
    # print(date_obj1,date_obj2)
    output = date_obj2 - date_obj1
    output = output.total_seconds()
    return output

def find_gap(traffic_data: list):
    output = []
    for index, entry in enumerate(traffic_data):
        bound = entry[0]
        if index+1 >= len(traffic_data):
            
            break
        for other_entry in traffic_data[index+1:]:
            if other_entry[0] == bound:
                # print(entry,other_entry)
                time_difference = find_time(entry,other_entry)
                break
        
        # print(target_time)
        # print(type(target_time))
        # time_difference = target_time - base_time
        # print(time_difference,entry[4])
        if time_difference % 1 == 0:
            time_difference = 0.5
        gap_distance = float(time_difference) * (float(entry[4]) /3.6)
        # if index == 2: 
        #     sys.exit()
        if gap_distance == 0:
            continue
        # print(gap_distance)
        output.append([entry, float(format(gap_distance,".2f"))])
        # output.append([entry[5], format(gap_distance,".2f")])
    # print(output)
    return output

def split_bound(traffic_data: list): 
    traffic_bound_1 = []
    traffic_bound_2 = []
    for entry in traffic_data:
        if entry == ['']:
            continue
        if int(entry[0]) == 1:
            traffic_bound_1.append(entry)
        else:
            traffic_bound_2.append(entry)
    return traffic_bound_1, traffic_bound_2

def split_lane(traffic_data: list):
    traffic_bound_1 = []
    traffic_bound_2 = []
    for entry in traffic_data:
        if entry == ['']:
            continue
        if int(entry[0]) == 1:
            traffic_bound_1.append(entry)
        else:
            traffic_bound_2.append(entry)
    return

def extract_data()-> list:
    file_names = ['tkb_wim-00aug23-1400 copy','tkb_wim-00aug23-2100 copy','tkb_wim-00nov21-1400 copy','tkb_wim-00nov21-2100 copy','tkb_wim-00nov21-2200 copy','tkb_wim-00nov22-0800 copy','tkb_wim-04dec30-0000 copy','tkb_wim-04dec30-0100 copy','tkb_wim-04dec30-0200 copy','tkb_wim-04dec30-0900 copy','tkb_wim-04dec30-1000 copy','tkb_wim-04dec30-1100 copy','tkb_wim-04dec30-1200 copy','tkb_wim-04dec30-1300 copy','tkb_wim-04dec30-1400 copy','tkb_wim-04dec30-1500 copy','tkb_wim-04dec30-1600 copy','tkb_wim-04dec30-1700 copy','tkb_wim-04dec30-1800 copy','tkb_wim-04dec30-1900 copy','tkb_wim-04dec30-2000 copy','tkb_wim-04dec30-2100 copy','tkb_wim-04dec30-2200 copy','tkb_wim-04dec30-2300 copy','tkb_win-00aug23-2100 copy']
    output = []
    traffic_data_sorted_bound = []
    for file_name in file_names: 
        print('now in: '+file_name)
        path = 'Final Year Project/WIM data/' + file_name + '.csv'
        with open(path,'r') as f: 
            txt = f.read()

        txt = txt.split("\n")

        # this is entirely diferent from extracting data
        for index, group in enumerate(txt):
            txt[index] = group.split(',')
        traffic_bound_1, traffic_bound_2 = split_bound(txt)
        # print(txt)
        traffic_data_sorted_bound = traffic_data_sorted_bound + traffic_bound_1 + traffic_bound_2
        output = output + find_gap(traffic_bound_1) + find_gap(traffic_bound_2)

    # print(len(output))
    # Output in the form of [traffic data sorted along sequence, gap distance]
    print(output[0])
    # print(len(output))
    return output

def add_vehicular_weight(output: list)-> list:
    for index, item in enumerate(output):
        # format for each item [['2', '23-AUG-2000 21:59:54', '1098', '1', '78', '2', '2', '310', '0', '310', '249'], '10.83']]
        vehicle_weight = 0
        vehicle_axle_weight_and_spacing_list = item[0][7:]
        for position in range(len(vehicle_axle_weight_and_spacing_list)):
            if position % 2 == 0:
                vehicle_weight += int(vehicle_axle_weight_and_spacing_list[position])
        # print(vehicle_weight)
        # print(output[index])
        # sys.exit()
        output[index].append(vehicle_weight)
    return output

def save_to_excel(output: list):
    my_wb = openpyxl.Workbook()
    my_sheet = my_wb.active

    heading = ['Gap Distance', 'Total Weight' ,'Bound',  'Date + Time',  'Seq No',  'Lane',  'Speed',  'Class', 'No of Axle', 'Axle Weight 1', 'Axle Spacing 1']
    for index_index, label in enumerate(heading):
        my_sheet.cell(row = 1, column = index_index + 1 ).value = label

    for index, item in enumerate(output): 
        # print(item)
        # sys.exit()
        index += 2
        my_sheet.cell(row = index, column = 1 ).value = item[1]
        my_sheet.cell(row = index, column = 2 ).value = item[2]
        for item_index, _ in enumerate(item[0]):
            column_position = item_index + 3
            my_sheet.cell(row = index, column = column_position ).value = item[0][item_index]

    my_wb.save('Final Year Project/output.xlsx')
    return

def plot_pdf(output: list):
    gap_distance_list = []
    weight_list = []
    largest_weight = [0,0,0]
    for item in output:
        if  (item[0][5] == '2' or item[0][5] == '1') and  (item[0][6] == '2' or item[0][6] == '1') :
            if int(item[2]) > largest_weight[2]:
                largest_weight = item
            gap_distance_list.append(float(item[1]))
            weight_list.append(float(item[2]))
    weight_list.sort()
    # for index, weight in enumerate(weight_list): 
    #     if int(weight) > 6000:
    #         weight_list = weight_list[:index]
    #         break
    # print(weight_list)

    #? Plots the Graph of GVW
    target_list = weight_list
    # Define the parameters of the normal distribution
    mean = statistics.mean(target_list)  # mean
    sigma = statistics.stdev(target_list)  # standard deviation
    print('mean and signma are: ' + str(mean) + ' '+ str(sigma))
    bins = 1000


    # generate random normal dataset
    _, bins, _ = plt.hist(target_list, 500, density=1, alpha=0.5)
    # mean, sigma = scipy.stats.norm.fit(target_list)
    best_fit_line = scipy.stats.norm.pdf(bins, mean, sigma)
    sns.distplot(target_list, hist=False, kde=True, rug = False, color = 'darkblue', hist_kws={'edgecolor':'black'},kde_kws={'linewidth': 4})
    plt.plot(bins, best_fit_line)
    plt.xlim([0,6000])
    plt.xlabel('Weight Distribution')
    plt.ylabel('Probability density')
    # plt.title('Class 3+ Gross Vehicle Weight Distribution')
    plt.title('Class 1 & 2 Gross Vehicle Weight Distribution')
    plt.show()
    sys.exit()

    # Generate random numbers from the normal distribution
    x = np.random.normal(mean, sigma, 10000)

    # Plot the PDF of the normal distribution                                               
    plt.hist(target_list, bins=bins, density=True, alpha=0.6, color='g')

    mu, sigma1 = scipy.stats.norm.fit(target_list)
    best_fit_line = scipy.stats.norm.pdf(bins, mu, sigma1)
    plt.plot(bins, best_fit_line)

    plt.xlim([0,25])
    plt.xlabel('Value')
    plt.ylabel('Probability density')
    plt.title('Normal distribution PDF')
    plt.show()
    return


    
output = extract_data()
# print(output[0])
output = add_vehicular_weight(output)
# plot_pdf(output)

save_to_excel(output)

# [0-9]+-[A-z]+-[0-9]+ +[0-9]+:[0-9]+:[0-9]+

# x_data = []
# y_data = []
# for item in output: 
#     if len(item[0]) <= 3: 
#         print(len(item[0]))
#         print(item[0])
#         sys.exit()
#     x_data.append(item[0][4])
#     y_data.append(item[1])
# print(x_data)
# print(y_data)

# #* Testing!!  
# d = {'speed': x_data, 'gap distance': y_data}
# df = pd.DataFrame(data=d)
 
# # use the function regplot to make a scatterplot
# sns.scatterplot(data= df, x= 'speed',y='gap distance')
 
# # make a scatterplot without regression fit
# #ax = sns.regplot(x=df["sepal_length"], y=df["sepal_width"], fit_reg=False)

# plt.show()
# sys.exit()



